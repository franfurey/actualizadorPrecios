import os
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from fuzzywuzzy import fuzz
import requests
import pandas as pd
from bs4 import BeautifulSoup
#import credentials
import re

############################################################################# CANAL ########################################################################################
############################################################################# CANAL ########################################################################################

def clean_canal(filename, new_filename=None,  mpn_value=None):
    # Cargar el archivo de CSV
    df = pd.read_csv(filename, encoding='ISO-8859-1', sep=';')


    df = df[['Identificador de URL', 'Nombre', 'Precio',
              'SKU','Código de barras','MPN (Número de pieza del fabricante)',
              'Costo','Categorías', 'Tags', 'Título para SEO', 'Descripción para SEO']]

    # Filtrar por valor específico en la columna MPN
    if mpn_value is not None:
        df = df[df['MPN (Número de pieza del fabricante)'] == mpn_value]

  # Limpiar la columna Precio
    df['Precio'] = df['Precio'].str.replace('\..*', '', regex=True).str.replace(',', '')
    df['Precio'] = pd.to_numeric(df['Precio'])


    # Extraer la palabra y el número de la columna SKU
    df[['SKU_palabra', 'SKU_num']] = df['SKU'].str.extract(r"(\D+)(\d+)")

    # Convertir la columna SKU_num a números
    df["SKU_num"] = pd.to_numeric(df["SKU_num"], errors="coerce")

    # Ordenar por SKU_num de forma ascendente
    df = df.sort_values(by='SKU_num')
    df['SKU_modificado'] = df['SKU_palabra'] + df['SKU_num'].astype(str)

    # Eliminar la columna SKU_letra
    df.drop(['SKU','SKU_num','SKU_palabra'], axis=1, inplace=True)
    df = df.rename(columns={'SKU_modificado': 'SKU'})

    # Guardar el archivo de Excel modificado en la carpeta "procesados"
    processed_dir = "./procesados/procesadosCanal"
    os.makedirs(processed_dir, exist_ok=True)  # crea la carpeta si no existe
    
    if new_filename is None:
        new_filename = os.path.splitext(os.path.basename(filename))[0] + '.xlsx'
    else:
        new_filename = new_filename + '.xlsx'
        
    processed_filename = os.path.join(processed_dir, new_filename)
    df.to_excel(processed_filename, index=False)
    
    print('Archivo limpio guardado como', processed_filename)

############################################################################# ALGABO ########################################################################################
############################################################################# ALGABO ########################################################################################

def clean_algabo(filename, new_filename=None):
    # Cargar el archivo de Excel
    df = pd.read_excel(filename)
    df = df.drop(index=range(16))

    df = df.rename(columns={
        df.columns[1]: 'SKU',
        df.columns[2]: 'Código de barras',
        df.columns[5]: 'Nombre',
        df.columns[26]: 'Costo'
    })
    df = df[['SKU', 'Código de barras', 'Nombre', 'Costo']]

    # Eliminar filas sin un código asociado
    columna = df.iloc[:, 1]
    df = df.dropna(subset=[df.columns[1]])
    df['Código de barras'] = df['Código de barras'].astype(int).astype(str)
    df['Código de barras'] = df['Código de barras'].str.replace(r'[.+\-E]', '', regex=True)
    df['Código de barras'] = df['Código de barras'].str.zfill(13)

    # Separar la columna SKU en dos columnas
    df["SKU"] = df["SKU"].astype(str)
    df[["num", "SKU_letra"]] = df["SKU"].str.extract(r"(\d+)(\D*)")

    # Convertir la columna SKU_numero a números
    df["num"] = pd.to_numeric(df["num"], errors="coerce")

    # Ordenar por SKU_numero de forma ascendente
    df = df.sort_values(by='num')

    # Concatenar SKU_numero y SKU_letra dentro de SKU_numero
    df['SKU'] =df['num'].astype(str) + df['SKU_letra']
    df = df.drop(columns=['num', 'SKU_letra'])

    # Aplicar un 15% de descuento a la columna Costo
    df['Costo'] = df['Costo'] * 0.85

    # Aplicar un 21% de aumento a la columna Costo
    df['Costo'] = df['Costo'] * 1.21

    # Guardar el archivo de Excel modificado en la carpeta "procesados"
    processed_dir = "./procesados"
    os.makedirs(processed_dir, exist_ok=True)  # crea la carpeta si no existe
    
    if new_filename is None:
        new_filename = os.path.basename(filename)
    else:
        _, file_extension = os.path.splitext(filename)
        new_filename = new_filename + file_extension
        
    processed_filename = os.path.join(processed_dir, new_filename)
    df.to_excel(processed_filename, index=False)

    print('Archivo limpio guardado como', processed_filename)

    # Eliminar las formas que no son imágenes
    wb = openpyxl.load_workbook(processed_filename)
    ws = wb.active

    # Iterar sobre todas las formas en la hoja de cálculo
    try:
        # Iterar sobre todas las formas en la hoja de cálculo
        for shape in ws._shapes:
            # Eliminar la forma si no es un gráfico
            if not isinstance(shape, openpyxl.drawing.image.Image):
                ws.remove_shape(shape)
    except AttributeError:
        # Si no se puede acceder a _shapes, usar _images en su lugar
        for image in ws._images:
            # Eliminar la imagen
            ws.remove_image(image)

    wb.save(processed_filename)

############################################################################# FUREY ########################################################################################
############################################################################# FUREY ########################################################################################

def clean_furey(filename, new_filename=None):
    # Cargar el archivo de Excel
    df = pd.read_excel(filename, dtype={'Código de barras': str})
    df = df.drop(index=range(6))

    df = df.rename(columns={
        df.columns[0]: 'SKU',
        df.columns[1]: 'Código de barras',
        df.columns[2]: 'Nombre',
        df.columns[3]: 'Costo'
    })
    df = df[['SKU', 'Código de barras', 'Nombre', 'Costo']]

    # Eliminar filas sin un código asociado
    columna = df.iloc[:, 0]
    df = df.dropna(subset=[df.columns[1]])
    df['Código de barras'] = df['Código de barras'].astype(str)
    df['Código de barras'] = df['Código de barras'].str.replace(r'[.+\-E]', '', regex=True)
    df['Código de barras'] = df['Código de barras'].str.zfill(13)

    # Separar la columna SKU en dos columnas
    df["SKU"] = df["SKU"].astype(str)
    df[["num", "SKU_letra"]] = df["SKU"].str.extract(r"(\d+)(\D*)")

    # Convertir la columna SKU_numero a números
    df["num"] = pd.to_numeric(df["num"], errors="coerce")

    # Ordenar por SKU_numero de forma ascendente
    df = df.sort_values(by='num')

    # Concatenar SKU_numero y SKU_letra dentro de SKU_numero
    df['SKU'] =df['num'].astype(str) + df['SKU_letra']
    df = df.drop(columns=['num', 'SKU_letra'])
    df['Costo'] = df['Costo'].astype(str)
    df['Costo'] = df['Costo'].str.replace(',', '').str.replace('\..*', '', regex=True)

    # DESCUENTOS 
    df['Costo'] = pd.to_numeric(df['Costo'])
    df['Costo'] = df['Costo'].apply(lambda x: round(x*0.83))


    # Guardar el archivo de Excel modificado en la carpeta "procesados"
    processed_dir = "./procesados"
    os.makedirs(processed_dir, exist_ok=True)  # crea la carpeta si no existe
    
    if new_filename is None:
        new_filename = os.path.basename(filename)
    else:
        _, file_extension = os.path.splitext(filename)
        new_filename = new_filename + file_extension
        
    processed_filename = os.path.join(processed_dir, new_filename)
    df.to_excel(processed_filename, index=False)

    print('Archivo limpio guardado como', processed_filename)

    # Eliminar las formas que no son imágenes
    wb = openpyxl.load_workbook(processed_filename)
    ws = wb.active

    # Iterar sobre todas las formas en la hoja de cálculo
    try:
        # Iterar sobre todas las formas en la hoja de cálculo
        for shape in ws._shapes:
            # Eliminar la forma si no es un gráfico
            if not isinstance(shape, openpyxl.drawing.image.Image):
                ws.remove_shape(shape)
    except AttributeError:
        # Si no se puede acceder a _shapes, usar _images en su lugar
        for image in ws._images:
            # Eliminar la imagen
            ws.remove_image(image)

    wb.save(processed_filename)

############################################################################# TEDDY ########################################################################################
############################################################################# TEDDY ########################################################################################

def clean_teddy(filename, new_filename=None):
    # Cargar el archivo de Excel
    df = pd.read_excel(filename)


    df = df.drop(index=range(19))

    df = df.rename(columns={
        df.columns[1]: 'SKU',
        df.columns[2]: 'Código de barras',
        df.columns[3]: 'Nombre',
        df.columns[4]: 'Costo'
    })
    df = df[['SKU', 'Código de barras', 'Nombre', 'Costo']]

    # Eliminar filas sin un código asociado
    columna = df.iloc[:, 1]
    df = df.dropna(subset=[df.columns[1]])
    df['Código de barras'] = df['Código de barras'].astype(int).astype(str)
    df['Código de barras'] = df['Código de barras'].str.replace(r'[.+\-E]', '', regex=True)
    df['Código de barras'] = df['Código de barras'].str.zfill(13)

    # Convertir la columna SKU_numero a números
    df["SKU"] = pd.to_numeric(df["SKU"], errors="coerce")

    # Ordenar por SKU_numero de forma ascendente
    df = df.sort_values(by='SKU')


    df['Costo'] = df['Costo'].replace(',', '', regex=True)

    # Limpiar los valores de la columna "Costo"
    df['Costo'] = df['Costo'].fillna(0).apply(lambda x: int(float(str(x).split('.')[0])))

    # Guardar el archivo de Excel modificado en la carpeta "procesados"
    processed_dir = "./procesados"
    os.makedirs(processed_dir, exist_ok=True)  # crea la carpeta si no existe
    
    if new_filename is None:
        new_filename = os.path.basename(filename)
    else:
        _, file_extension = os.path.splitext(filename)
        new_filename = new_filename + file_extension
        
    processed_filename = os.path.join(processed_dir, new_filename)
    df.to_excel(processed_filename, index=False)

    print('Archivo limpio guardado como', processed_filename)

    # Eliminar las formas que no son imágenes
    wb = openpyxl.load_workbook(processed_filename)
    ws = wb.active

    # Iterar sobre todas las formas en la hoja de cálculo
    try:
        # Iterar sobre todas las formas en la hoja de cálculo
        for shape in ws._shapes:
            # Eliminar la forma si no es un gráfico
            if not isinstance(shape, openpyxl.drawing.image.Image):
                ws.remove_shape(shape)
    except AttributeError:
        # Si no se puede acceder a _shapes, usar _images en su lugar
        for image in ws._images:
            # Eliminar la imagen
            ws.remove_image(image)

            #

    wb.save(processed_filename)
    print('Imágenes eliminadas del archivo', processed_filename)

    ############################################################################# DRIMEL ########################################################################################
    ############################################################################# DRIMEL ########################################################################################

def scrape_drimel(urls):
    session = requests.Session()

    marcas = ['Algabo', 'Duffy', 'Huggies','Babysec','Caricia','Estrella','Pampers',
              'Candy','Doncella','Deyse','Johnsons','Kimbies','Upa']

    df = pd.DataFrame(columns=['name', 'price','marca','product_id', 'product_sku','url_id','MPN (Número de pieza del fabricante)'])        
    
    
    # Buscamos los elementos que contienen la información de los productos
    # https://drimel.com.ar/?product_cat=panales
    # https://drimel.com.ar/?product_cat=panales&paged=2
    for extract_url in urls: 
        page = 1
        while True:
            url = extract_url + '&paged={}'.format(page)

            response = requests.get(url)
            soup = BeautifulSoup(response.content, 'html.parser')

            products = soup.find_all('div', class_='box-text box-text-products')
            #print('Products',products)
            if not products:
                break

            # Iteramos sobre los elementos encontrados y extraemos la información deseada
            for product in products:
                
                name = product.find('a', class_='woocommerce-LoopProduct-link').text.strip()
            
                price = product.find('span', class_='woocommerce-Price-amount').text.strip()[1:] # Quitamos el signo $
                price = price.replace('"', '').replace('.', '').split(',')[0] # Quitamos las comillas, la coma y lo que hay después del punto
                
                
                # Buscamos si el nombre del producto contiene alguna de las marcas
                marca = ''
                for m in marcas:
                    if m.lower() in name.lower():
                        marca = m
                        break
                # Obtenemos el código de barras y el código común del producto
                add_to_cart_button = product.find('a', class_='add_to_cart_button')
                product_id = add_to_cart_button.get('data-product_id', '')
                product_sku = add_to_cart_button.get('data-product_sku', '')

                url_id = re.sub('[^a-zA-Z0-9]+', '-', name).strip('-')
                url_id = url_id.replace(' ', '-')

                # Excluimos las filas que tengan como marca 'Algabo' o 'Upa'
                df = df[(df['marca'] != 'Algabo') & (df['marca'] != 'Upa')]

                MPN = 'Drimel'
                df.loc[len(df)] = [name, price, marca, product_id, product_sku, url_id, MPN]
            page += 1
                # Puedes hacer lo que necesites con la información obtenida, como imprimir o guardar en un archivo
    df = df.rename(columns={'name': 'Nombre', 'price': 'Costo',
                             'product_id': 'SKU', 'product_sku': 'Código de barras',
                               'url_id':'Identificador de URL', 'marca':'Marca'})
    

    df['Costo'] = pd.to_numeric(df['Costo'])
    df['Costo'] = df['Costo'].apply(lambda x: round(x*0.9))



    df['SKU'] = df['SKU'].astype(int)
    df = df.sort_values('SKU', ascending=True)

    df.to_excel('./procesados/drimel.xlsx', index=False)
    print('Archivo guardado con exito')
    return

############################################################################# UPALALA ########################################################################################
############################################################################# UPALALA ########################################################################################

def clean_upalala(filename, new_filename=None):
    # Cargar el archivo de Excel
    df = pd.read_excel(filename)
    df = df.drop(index=range(0))

    df = df.rename(columns={
        df.columns[2]: 'SKU',
        df.columns[0]: 'Código de barras',
        df.columns[3]: 'Nombre',
        df.columns[4]: 'Costo'
    })
    df = df[['SKU', 'Código de barras', 'Nombre', 'Costo']]

    # Eliminar filas sin un código asociado
    columna = df.iloc[:, 1]
    df = df.dropna(subset=[df.columns[1]])
    df['Código de barras'] = df['Código de barras'].astype(int).astype(str)
    df['Código de barras'] = df['Código de barras'].str.replace(r'[.+\-E]', '', regex=True)
    df['Código de barras'] = df['Código de barras'].str.zfill(13)

    df['Costo'] = df['Costo'].astype(str).str.split('.', n=1, expand=True)[0]



    # Guardar el archivo de Excel modificado en la carpeta "procesados"
    processed_dir = "./procesados"
    os.makedirs(processed_dir, exist_ok=True)  # crea la carpeta si no existe
    
    if new_filename is None:
        new_filename = os.path.basename(filename)
    else:
        _, file_extension = os.path.splitext(filename)
        new_filename = new_filename + file_extension
        
    processed_filename = os.path.join(processed_dir, new_filename)
    df.to_excel(processed_filename, index=False)

    print('Archivo limpio guardado como', processed_filename)

    # Eliminar las formas que no son imágenes
    wb = openpyxl.load_workbook(processed_filename)
    ws = wb.active

    # Iterar sobre todas las formas en la hoja de cálculo
    try:
        # Iterar sobre todas las formas en la hoja de cálculo
        for shape in ws._shapes:
            # Eliminar la forma si no es un gráfico
            if not isinstance(shape, openpyxl.drawing.image.Image):
                ws.remove_shape(shape)
    except AttributeError:
        # Si no se puede acceder a _shapes, usar _images en su lugar
        for image in ws._images:
            # Eliminar la imagen
            ws.remove_image(image)

    wb.save(processed_filename)