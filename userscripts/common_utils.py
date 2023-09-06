# userscripts/common_utils.py
import os
import json
import openpyxl
import traceback
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter




#

def eliminar_formas_no_imagenes(filename):
    ext = os.path.splitext(filename)[1]
    if ext not in ['.xlsx', '.xlsm']:
        print(f"Omitiendo {filename}, ya que no es un archivo Excel soportado.")
        return
    
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    try:
        for shape in ws._shapes:
            if not isinstance(shape, openpyxl.drawing.image.Image):
                ws.remove_shape(shape)
    except AttributeError:
        for image in ws._images:
            ws.remove_image(image)
    wb.save(filename)

#

def adjust_columns_and_center_text(path):
    try:
        # Obtenemos la lista de archivos en el directorio
        files = os.listdir(path)
        print(f"Archivos en el directorio {path}: {files}")

        for file in files:
            # Procesamos solo los archivos .xlsx
            if file.endswith('.xlsx'):
                print(f"Procesando archivo: {file}")
                try:
                    # Abrimos el archivo usando openpyxl
                    workbook = openpyxl.load_workbook(os.path.join(path, file))

                    # Iteramos a través de todas las hojas en el archivo
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]

                        # Iteramos a través de todas las columnas en la hoja
                        for col in sheet.columns:
                            max_length = 0
                            column = col[0].column_letter  # Obtenemos la letra de la columna

                            # Encontramos el máximo ancho de la columna
                            for cell in col:
                                try:
                                    max_length = max(max_length, len(str(cell.value)))
                                    # Centramos el texto en la celda
                                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                                except:
                                    print("Error al ajustar celda:", cell)
                                    pass

                            # Ajustamos el ancho de la columna
                            adjusted_width = (max_length + 2)
                            sheet.column_dimensions[column].width = adjusted_width

                    # Guardamos los cambios en el archivo
                    workbook.save(os.path.join(path, file))
                    print(f"Archivo ajustado y centrado: {file}")
                except Exception as e:
                    print("Error en el archivo:", file)
                    traceback.print_exc()
    except Exception as e:
        print("Error en adjust_columns_and_center_text:")
        traceback.print_exc()

#

def generate_report_and_pdf(df_final, path, proveedor_nombre, df_hoja1, df_hoja2, df_hoja3, porcentaje_aumento, user):
    total_rows = len(df_final)
    matched_products = len(df_hoja1)
    missing_df1_rows = len(df_hoja2)
    missing_df2_rows = len(df_hoja3)
    # Calculo de productos con aumento y su porcentaje promedio
    increased_products = df_hoja1[df_hoja1['Porcentaje_Aumento'] > 0]
    if increased_products.empty:
        avg_increase_percent = 0
    else:
        avg_increase_percent = round(increased_products['Porcentaje_Aumento'].mean())

    # Calculo de productos con descuento y su porcentaje promedio
    discounted_products = df_hoja1[df_hoja1['Porcentaje_Aumento'] < 0]
    if discounted_products.empty:
        avg_discount_percent = 0
    else:
        avg_discount_percent = round(discounted_products['Porcentaje_Aumento'].mean())


    report_data = {
        "matched_products": matched_products,
        "total_products_supplier_db": total_rows,
        "missing_df1_rows": missing_df1_rows,
        "missing_df2_rows": missing_df2_rows,
        "avg_increase_percent": avg_increase_percent,
        "increased_products_count": len(increased_products),
        "avg_discount_percent": avg_discount_percent,
        "discounted_products_count": len(discounted_products),
        "porcentaje_aumento_cliente": porcentaje_aumento
    }

    print("Datos del informe antes de escribir en el archivo JSON:", report_data)
    with open(os.path.join(path, "report_data.json"), "w") as json_file:
        json.dump(report_data, json_file)

    pdf_path = os.path.join(path, f"{proveedor_nombre}-Reporte-PS.pdf")
    c = canvas.Canvas(pdf_path, pagesize=letter)

    c.setFont("Helvetica-Bold", 18)
    c.drawString(30, 730, f"Informe de Actualizacion de precios de {proveedor_nombre} para {user.company}")  # Título bajado a 730
    c.setFont("Helvetica", 12)
    c.drawString(30, 700, f"Margen marcado por {user.company}: {porcentaje_aumento}%")  # Bajado a 700
    c.drawString(30, 680, f"Total de productos encontrados en ambas listas: {matched_products}")  # Bajado a 680
    c.drawString(30, 660, f"Total de productos que están en la base de datos de {user.company} pero no en la de {proveedor_nombre}: {missing_df1_rows}")  # Bajado a 660
    c.drawString(30, 640, f"Productos nuevos incorporados por {proveedor_nombre} pero no incorporados en {user.company}: {missing_df2_rows}")  # Bajado a 640
    if avg_increase_percent is not None and increased_products is not None:
        c.drawString(30, 620, f"Cantidad de productos con aumento: {len(increased_products)}, con un promedio de aumento del {int(avg_increase_percent)}%")  # Bajado a 620
    if avg_discount_percent is not None and discounted_products is not None:
        c.drawString(30, 600, f"Cantidad de productos con descuento: {len(discounted_products)}, con un promedio de descuento del {int(avg_discount_percent)}%")  # Bajado a 600
    c.save()
    print(f"Archivo de informe guardado como {pdf_path}")


    return report_data