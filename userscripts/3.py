# userscripts/3.py
from reportlab.lib.pagesizes import letter
from fuzzywuzzy import fuzz
from reportlab.pdfgen import canvas
from openpyxl import load_workbook
import importlib
import os
import pandas as pd
import openpyxl as px
import traceback
import json

def rename_files(current_user_id, proveedor, path):
    # Obtenemos la lista de archivos en el directorio.
    files = os.listdir(path)
    print(f"Archivos en el directorio {path}: {files}")

    for file in files:
        # Procesamos solo los archivos .xls y .xlsx
        if file.endswith('.xls') or file.endswith('.xlsx'):
            print(f"Procesando archivo: {file}")
            try:
                # Leemos las primeras 5 filas del archivo para determinar su tipo.
                df = pd.read_excel(os.path.join(path, file), nrows=5)
                # Obtenemos la extensión del archivo original
                file_extension = os.path.splitext(file)[1]
                # Dependiendo del contenido, renombramos el archivo usando la extensión original
                if 'Gestión Pedidos/Reposición de Stock' in df.to_string():
                    new_name = f'{current_user_id}-{proveedor.nombre}-stock{file_extension}'
                elif 'Articulos' in df.to_string():
                    new_name = f'{current_user_id}-{proveedor.nombre}-precios{file_extension}'
                elif 'Precios de Listas de Proveedores' in df.to_string():
                    new_name = f'{current_user_id}-{proveedor.nombre}-codProveedor{file_extension}'
                else:
                    new_name = f'{current_user_id}-{proveedor.nombre}-proveedor{file_extension}'
                # Renombramos el archivo.
                os.rename(os.path.join(path, file), os.path.join(path, new_name))
                print(f"Archivo renombrado a: {new_name}")
            except Exception as e:
                print("Error en rename_files:")
                traceback.print_exc()



########################################################################################################################################################################

def leer_archivo(filename, skiprows=0, sku_as_str=False):
    """
    Leer un archivo CSV, XLS o XLSX y devuelve un DataFrame.
    La función detecta automáticamente el tipo de archivo y realiza la lectura correspondiente.
    """
    try:
        if filename.endswith('.csv'):
            df = pd.read_csv(filename, encoding='ISO-8859-1', sep=';', skiprows=skiprows)
        elif filename.endswith('.xls'):
            if sku_as_str:
                df = pd.read_excel(filename, skiprows=skiprows, converters={2: str})
            else:
                df = pd.read_excel(filename, skiprows=skiprows)
            new_filename = filename[:-4] + '.xlsx'  # Crear el nombre del nuevo archivo
            df.to_excel(new_filename, index=False)  # Guardar el DataFrame como .xlsx
            os.remove(filename)  # Eliminar el archivo .xls original
            filename = new_filename  # Actualizar el nombre del archivo para las operaciones siguientes
        elif filename.endswith('.xlsx'):
            book = load_workbook(filename, data_only=True)
            sheet = book.active
            data = sheet.values
            columns = next(data)[skiprows:]
            df = pd.DataFrame(data, columns=columns)
            if sku_as_str:
                df[2] = df[2].astype(str)
        else:
            print(f"Formato de archivo no soportado: {filename}")
            return None, None
    except pd.errors.EmptyDataError:
        print(f"El archivo está vacío: {filename}")
        return None, None
    except FileNotFoundError:
        print(f"El archivo no existe: {filename}")
        return None, None
    except Exception as e:
        print(f"Error al leer el archivo: {filename}, {str(e)}")
        return None, None
     # Agregar este fragmento justo después de leer el archivo y antes de devolver los resultados
    if df is not None:
        # Elimina las columnas que tengan 'None' como nombre
        df = df.drop(columns=[col for col in df.columns if col is None])
    
    return df, filename  # Devolver el DataFrame y el nombre del archivo

########################################################################################################################################################################

def eliminar_formas_no_imagenes(filename):
    wb = px.load_workbook(filename)
    ws = wb.active
    try:
        for shape in ws._shapes:
            if not isinstance(shape, px.drawing.image.Image):
                ws.remove_shape(shape)
    except AttributeError:
        for image in ws._images:
            ws.remove_image(image)
    wb.save(filename)

########################################################################################################################################################################

def convert_date_to_code(value):
    if '-' in value and len(value.split('-')) == 3:
        parts = value.split('-')
        return f"{parts[1]}-{parts[0]}"
    return value


########################################################################################################################################################################

def filtrar_y_reformatear_mg_precio(df):
    """Filtrar y reformatear un DataFrame"""
    # Obtener los nombres de las columnas de la fila con índice 2
    # Eliminar las primeras tres filas y asignar los nombres de las columnas
    df = df.copy()
    column_names = df.iloc[2].values
    df = df.iloc[3:]
    df.columns = column_names
    
    # Cambiar el nombre de la columna 'Código Alt.' a 'SKU'
    df.rename(columns={'Código Alt.': 'SKU'}, inplace=True)
    df.rename(columns={'Precio 1': 'Precio'}, inplace=True)
    df['Precio'] = df['Precio'].apply(lambda x: int(x))
    df.rename(columns={'Descripción del Artículo': 'Nombre'}, inplace=True)
    # Cambiar el nombre de la columna 'Código' a 'Codigo'
    df.rename(columns={'Código': 'Codigo'}, inplace=True)
    # Convierte la columna 'SKU' a string, eliminando los espacios y manejando los valores nulos
    df['SKU'] = df['SKU'].apply(lambda x: '{:.0f}'.format(float(str(x).replace(" ", ""))) if pd.notnull(x) else 'NaN')

    return df


########################################################################################################################################################################

def filtrar_y_reformatear_mg_stock(df):
    """Filtrar y reformatear un DataFrame"""
    # Obtener los nombres de las columnas de la fila con índice 2
    # Eliminar las primeras tres filas y asignar los nombres de las columnas
    df = df.copy()
    column_names = df.iloc[2].values
    df = df.iloc[3:]
    df.columns = column_names
    # Cambiar el nombre de la columna 'Código' a 'Codigo'
    df.rename(columns={'Código': 'Codigo'}, inplace=True)
    df.rename(columns={'Descripción': 'Nombre'}, inplace=True)
    # Restablecer el índice para que comience desde 0
    df.reset_index(drop=True, inplace=True)

    return df


########################################################################################################################################################################

def filtrar_y_reformatear_codProveedor (df):
    """Filtrar y reformatear un DataFrame"""
    # Obtener los nombres de las columnas de la fila con índice 2
    # Eliminar las primeras tres filas y asignar los nombres de las columnas
    df = df.copy()
    column_names = df.iloc[2].values
    df = df.iloc[3:]
    df.columns = column_names
    df.rename(columns={'Cód.de Artículo': 'Codigo'}, inplace=True)
    df.rename(columns={'Descripción del Artículo': 'Nombre'}, inplace=True)
    df.rename(columns={'Cód.Art.según Proveedor': 'Codigo_segun_Proveedor'}, inplace=True)
    df.reset_index(drop=True, inplace=True)
    return df


########################################################################################################################################################################

def combine_dataframes(df1, df2, suffix='_df2'):
    df2.columns = [col + suffix if col in df1.columns else col for col in df2.columns]
    rows_list = []

    for i, row1 in df1.iterrows():
        for j, row2 in df2.iterrows():
            if row1['Codigo'] == row2['Codigo' + suffix]: # Comparación exacta aquí
                combined_row = pd.concat([row1, row2])
                rows_list.append(combined_row)
                break
        else:
            rows_list.append(row1)

    for i, row2 in df2.iterrows():
        if not any(row2['Codigo' + suffix] == row1['Codigo'] for _, row1 in df1.iterrows()): # Comparación exacta aquí
            rows_list.append(row2)

    result_df = pd.DataFrame(rows_list)
    return result_df



########################################################################################################################################################################

def combine_dataframes_sku(df1, df2):
    df2.columns = [col + '_proveedor' if col in df1.columns else col for col in df2.columns]
    rows_list = []
    df1['similarity'] = 0

    for i, row1 in df1.iterrows():
        match_found = False
        for j, row2 in df2.iterrows():
            similarity = 0
            code_match = False
            
            if 'SKU' in row1 and 'SKU_proveedor' in row2:
                similarity = fuzz.token_set_ratio(row1['SKU'], row2['SKU_proveedor'])
            
            if 'Codigo_proveedor' in row2 and 'Codigo_segun_Proveedor' in row1:
                code_match = row1['Codigo_segun_Proveedor'] == row2['Codigo_proveedor']

            if similarity >= 94 or code_match:
                combined_row = pd.concat([row1, row2])
                combined_row['similarity'] = similarity if similarity >= 94 else 55  # Usar 55 si es un código coincidente
                rows_list.append(combined_row)
                match_found = True
                break

        if not match_found:
            rows_list.append(row1)

    for i, row2 in df2.iterrows():
        if not any(('SKU_proveedor' in row2 and fuzz.token_set_ratio(row2['SKU_proveedor'], row1['SKU']) >= 94) or
                   ('Codigo_proveedor' in row2 and 'Codigo_segun_Proveedor' in row1 and row2['Codigo_proveedor'] == row1['Codigo_segun_Proveedor']) for _, row1 in df1.iterrows()):
            rows_list.append(row2)

    result_df = pd.DataFrame(rows_list)
    return result_df


########################################################################################################################################################################

def seleccionar_columnas(df_final, proveedor, path):
    def eliminar_filas_vacias(df):
        return df.dropna(how='all')

    file_name = f"{proveedor.nombre}.xlsx"
    writer = pd.ExcelWriter(os.path.join(path, file_name))

    columnas_hoja1 = ["Codigo_proveedor", "Codigo", "SKU", "Nombre", "Stock Act.", "Stock Max.", "Pto. Repos.", "Costo", "Precio"]
    columnas_hoja2 = ["Codigo", "Codigo_segun_Proveedor", "SKU", "Nombre", "Stock Act.", "Stock Max.", "Pto. Repos.", "Precio"]
    columnas_hoja3 = ["Codigo_proveedor", "SKU_proveedor", "Nombre_proveedor", "Costo"]

    columnas_hoja1 = [col for col in columnas_hoja1 if col in df_final.columns]
    columnas_hoja2 = [col for col in columnas_hoja2 if col in df_final.columns]
    columnas_hoja3 = [col for col in columnas_hoja3 if col in df_final.columns]

    df_hoja1 = eliminar_filas_vacias(df_final[df_final['similarity'] > 50][columnas_hoja1])
    df_hoja1.to_excel(writer, sheet_name='Productos en Común', index=False)

    if 'SKU_proveedor' in df_final.columns:
        filter_condition_2 = df_final['SKU_proveedor'].isnull()
    elif 'Codigo_proveedor' in df_final.columns:
        filter_condition_2 = df_final['Codigo_proveedor'].isnull()
    else:
        filter_condition_2 = False

    df_hoja2 = eliminar_filas_vacias(df_final[filter_condition_2][columnas_hoja2])
    df_hoja2 = df_hoja2.sort_values(by="Stock Act.", ascending=False)
    df_hoja2.to_excel(writer, sheet_name='Productos Descontinuados', index=False)

    if 'SKU' in df_final.columns:
        filter_condition_3 = df_final['SKU'].isnull()
    elif 'Codigo' in df_final.columns:
        filter_condition_3 = df_final['Codigo'].isnull()
    else:
        filter_condition_3 = False

    df_hoja3 = eliminar_filas_vacias(df_final[filter_condition_3][columnas_hoja3])
    df_hoja3.to_excel(writer, sheet_name='Nuevos Productos', index=False)

    writer.save()


########################################################################################################################################################################

def adjust_columns_and_center_text(path):
    try:
        # Obtenemos la lista de archivos en el directorio
        files = os.listdir(path)
        print(f"Archivos en el directorio {path}: {files}")

        for file in files:
            # Procesamos solo los archivos .xlsx
            if file.endswith('.xlsx'):
                print(f"Procesando archivo: {file}")
                try:
                    # Abrimos el archivo usando openpyxl
                    workbook = px.load_workbook(os.path.join(path, file))

                    # Iteramos a través de todas las hojas en el archivo
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]

                        # Iteramos a través de todas las columnas en la hoja
                        for col in sheet.columns:
                            max_length = 0
                            column = col[0].column_letter  # Obtenemos la letra de la columna

                            # Encontramos el máximo ancho de la columna
                            for cell in col:
                                try:
                                    max_length = max(max_length, len(str(cell.value)))
                                    # Centramos el texto en la celda
                                    cell.alignment = px.styles.Alignment(horizontal='center')
                                except:
                                    print("Error al ajustar celda:", cell)
                                    pass

                            # Ajustamos el ancho de la columna
                            adjusted_width = (max_length + 2)
                            sheet.column_dimensions[column].width = adjusted_width

                    # Guardamos los cambios en el archivo
                    workbook.save(os.path.join(path, file))
                    print(f"Archivo ajustado y centrado: {file}")
                except Exception as e:
                    print("Error en el archivo:", file)
                    traceback.print_exc()
    except Exception as e:
        print("Error en adjust_columns_and_center_text:")
        traceback.print_exc()


########################################################################################################################################################################

def process_files(current_user_id, proveedor, path):
    # 1. Renombrar archivos
    rename_files(current_user_id, proveedor, path)
    
    # Inicializar los dataframes como None
    df_precios = None
    df_stock = None
    df_codProveedor = None
    df_proveedor = None

    # Leer y procesar los archivos
    files = os.listdir(path)
    print(f"Lista de archivos después de renombrar: {files}")  # Añade esta línea

    for file in files:
        if file.endswith('.xls') or file.endswith('.xlsx'):
            print(f"Procesando archivo: {file}")
            try:
                if 'stock' in file or 'precios' in file:
                    df, filename = leer_archivo(os.path.join(path, file), skiprows=0, sku_as_str='precios' in file)
                else:
                    df, filename = leer_archivo(os.path.join(path, file))

                # 3. Eliminar formas no imágenes
                eliminar_formas_no_imagenes(filename)

                # 4. Filtrar y reformatear el DataFrame
                if '-precios.xlsx' in os.path.basename(filename):
                    df_precios = filtrar_y_reformatear_mg_precio(df)
                    df_precios.to_excel(filename, index=False)
                elif '-stock.xlsx' in os.path.basename(filename):
                    df_stock = filtrar_y_reformatear_mg_stock(df)
                    df_stock.to_excel(filename, index=False)
                elif '-codProveedor.xlsx' in os.path.basename(filename):
                    # Aquí puedes llamar a una función para filtrar y reformatear el DataFrame
                    # de la manera que necesites para el archivo -codProveedor.
                    # Por ejemplo:
                    df_codProveedor = filtrar_y_reformatear_codProveedor(df)
                    df_codProveedor.to_excel(filename, index=False)
                # Resto del código para procesar otros archivos ...


                # Procesar archivos de proveedor
                if '-proveedor.xlsx' in os.path.basename(filename):
                    # Importar el script del proveedor
                    script_proveedor = importlib.import_module(f'proveedores.{current_user_id}.{proveedor.id}')

                    # Añade la impresión del DataFrame aquí antes de la llamada a process_proveedor_file
                    print(f"DataFrame antes de process_proveedor_file para el archivo {file}:")
                    print(df.head(10))

                    # Llamar a la función process_proveedor_file en el script del proveedor
                    df_proveedor = script_proveedor.process_proveedor_file(df)  # Cambiado a df_proveedor

                    # Añade la impresión del DataFrame aquí
                    print(f"DataFrame después de process_proveedor_file para el archivo {file}:")
                    print(df_proveedor.head(10))
                    df_proveedor.to_excel(filename, index=False)  # Cambiado a df_proveedor
            except Exception as e:
                print(f"Error al procesar el archivo {file}:")
                traceback.print_exc()

    # Definir df_combined antes de las condicionales
    df_combined = None

    # Combinar los dataframes de precios y stock
    if df_precios is not None and df_stock is not None:
        df_combined = combine_dataframes(df_precios, df_stock)
        df_combined.to_excel(os.path.join(path, "combined.xlsx"), index=False)
    else:
        print("No se encontraron los archivos de precios y stock.")
    # Combinar el dataframe combinado con el dataframe de codProveedor (si existe)
    if df_combined is not None and df_codProveedor is not None:
        df_combined = combine_dataframes(df_combined, df_codProveedor, suffix='_df3')
        df_combined.to_excel(os.path.join(path, "combined.xlsx"), index=False)

    # Combinar el dataframe combinado con el dataframe de proveedor
    if df_combined is not None and df_proveedor is not None:
        
        df_final = combine_dataframes_sku(df_combined, df_proveedor)
        # Llamar a la función para generar el informe
        # generate_report_and_pdf(df_final, path, proveedor.nombre)
        df_final.to_excel(os.path.join(path, "final.xlsx"), index=False)
        # Crear un nuevo archivo Excel con las hojas y columnas seleccionadas
        seleccionar_columnas(df_final, proveedor, path)
        adjust_columns_and_center_text(path)
    else:
        print("No se encontraron los archivos combinados y de proveedor.")