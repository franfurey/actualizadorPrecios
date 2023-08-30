# userscripts/3.py
# Este es el archivo de fran@soycanal.com.ar
from reportlab.lib.pagesizes import letter
from fuzzywuzzy import fuzz
from reportlab.pdfgen import canvas
from openpyxl import load_workbook
import importlib
import os
import pandas as pd
import openpyxl as px
import traceback
import json
import openpyxl

def rename_files(current_user_id, proveedor, path):
    files = os.listdir(path)
    for file in files:
        if file.endswith('.csv') or file.endswith('.xls') or file.endswith('.xlsx'):
            new_name = f"{current_user_id}-{proveedor.nombre}-{'canal' if 'canal' in file else 'proveedor'}.xlsx"
            os.rename(os.path.join(path, file), os.path.join(path, new_name))
            print(f"Archivo renombrado a: {new_name}")

########################################################################################################################################################################            

def leer_archivo(filename, skiprows=0):
    try:
        df = None
        if filename.endswith('.csv'):
            df = pd.read_csv(filename, encoding='ISO-8859-1', sep=';')
        elif filename.endswith('.xls') or filename.endswith('.xlsx'):
            df = pd.read_excel(filename, skiprows=skiprows)

        if df is not None:
            new_filename = filename.split('.')[0] + '.xlsx'
            df.to_excel(new_filename, index=False)
            if filename != new_filename:
                os.remove(filename)
            
            # Elimina columnas con 'None' como nombre
            df.drop(columns=[col for col in df.columns if col is None], inplace=True)
            
        return df, new_filename

    except Exception as e:
        print(f"Error al leer el archivo: {filename}, {str(e)}")
        traceback.print_exc()
        return None, None

########################################################################################################################################################################            

def eliminar_formas_no_imagenes(filename):
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    try:
        for shape in ws._shapes:
            if not isinstance(shape, openpyxl.drawing.image.Image):
                ws.remove_shape(shape)
    except AttributeError:
        for image in ws._images:
            ws.remove_image(image)
    wb.save(filename)

########################################################################################################################################################################

def filtrar_y_reformatear_canal(df, mpn_value=None):
    """Filtrar y reformatear un DataFrame"""
    df = df[['Identificador de URL', 'Nombre', 'Precio', 'SKU','Código de barras','MPN (Número de pieza del fabricante)', 'Costo','Categorías', 'Tags', 'Título para SEO', 'Descripción para SEO']]
    if mpn_value is not None:
        df = df[df['MPN (Número de pieza del fabricante)'] == mpn_value]

    # Limpiar la columna Precio
    df['Precio'] = df['Precio'].str.replace('\..*', '', regex=True).str.replace(',', '')
    df['Precio'] = pd.to_numeric(df['Precio'])

    # Extraer la palabra y el número de la columna SKU
    df[['SKU_palabra', 'SKU_num']] = df['SKU'].str.extract(r"(\D+)(\d+)")

    # Convertir la columna SKU_num a números
    df["SKU_num"] = pd.to_numeric(df["SKU_num"], errors="coerce")

    # Ordenar por SKU_num de forma ascendente
    df = df.sort_values(by='SKU_num')
    df['SKU_modificado'] = df['SKU_palabra'] + df['SKU_num'].astype(str)

    # Eliminar la columna SKU_letra
    df.drop(['SKU','SKU_num','SKU_palabra'], axis=1, inplace=True)
    df = df.rename(columns={'SKU_modificado': 'SKU'})
    return df

########################################################################################################################################################################

def combine_dataframes(df1, df2):
    df1 = df1.copy()  # Copiar df1 para evitar cambios en el original
    df2 = df2.copy()  # Copiar df2 para evitar cambios en el original
    rows_list = []
    
    # Ajustar nombres de columnas en df2
    df2.columns = [col + '_proveedor' if col in df1.columns else col for col in df2.columns]
    
    for i, row1 in df1.iterrows():
        match_found = False
        for j, row2 in df2.iterrows():
            # Inicializar variables
            similarity = 0
            code_match = False
            
            # Calcular similitud en "Código de barras" si está disponible
            if 'Código de barras' in row1 and 'Código de barras_proveedor' in row2:
                similarity = fuzz.token_set_ratio(row1['Código de barras'], row2['Código de barras_proveedor'])

            # Verificar coincidencia de códigos si están disponibles
            if 'SKU_proveedor' in row2 and 'SKU' in row1:
                code_match = row1['SKU'] == row2['SKU_proveedor']
            
            if similarity >= 94 or code_match:
                combined_row = pd.concat([row1, row2])
                combined_row['similarity'] = similarity if similarity >= 94 else 55
                rows_list.append(combined_row)
                match_found = True
                break

        if not match_found:
            rows_list.append(row1)

    # Añadir filas de df2 que no coinciden con ninguna fila de df1
    for i, row2 in df2.iterrows():
        if not any((fuzz.token_set_ratio(row2['Código de barras_proveedor'], row1['Código de barras']) >= 94) or (row2['SKU_proveedor'] == row1['SKU']) for _, row1 in df1.iterrows()):
            rows_list.append(row2)

    result_df = pd.DataFrame(rows_list)
    return result_df

########################################################################################################################################################################

def seleccionar_columnas(df_final, proveedor, path):
    def eliminar_filas_vacias(df):
        return df.dropna(how='all')

    file_name = f"{proveedor.nombre}.xlsx"
    writer = pd.ExcelWriter(os.path.join(path, file_name))

    columnas_hoja1 = ["Codigo_proveedor", "Codigo", "SKU", "Nombre", "Stock Act.", "Stock Max.", "Pto. Repos.", "Costo", "Precio"]
    columnas_hoja2 = ["Codigo", "Codigo_segun_Proveedor", "SKU", "Nombre", "Stock Act.", "Stock Max.", "Pto. Repos.", "Precio"]
    columnas_hoja3 = ["Codigo_proveedor", "SKU_proveedor", "Nombre_proveedor", "Costo"]

    columnas_hoja1 = [col for col in columnas_hoja1 if col in df_final.columns]
    columnas_hoja2 = [col for col in columnas_hoja2 if col in df_final.columns]
    columnas_hoja3 = [col for col in columnas_hoja3 if col in df_final.columns]

    df_hoja1 = eliminar_filas_vacias(df_final[df_final['similarity'] > 50][columnas_hoja1])
    df_hoja1.to_excel(writer, sheet_name='Productos en Común', index=False)

    if 'SKU_proveedor' in df_final.columns:
        filter_condition_2 = df_final['SKU_proveedor'].isnull()
    elif 'Codigo_proveedor' in df_final.columns:
        filter_condition_2 = df_final['Codigo_proveedor'].isnull()
    else:
        filter_condition_2 = False

    df_hoja2 = eliminar_filas_vacias(df_final[filter_condition_2][columnas_hoja2])
    df_hoja2 = df_hoja2.sort_values(by="Stock Act.", ascending=False)
    df_hoja2.to_excel(writer, sheet_name='Productos Descontinuados', index=False)

    if 'SKU' in df_final.columns:
        filter_condition_3 = df_final['SKU'].isnull()
    elif 'Codigo' in df_final.columns:
        filter_condition_3 = df_final['Codigo'].isnull()
    else:
        filter_condition_3 = False

    df_hoja3 = eliminar_filas_vacias(df_final[filter_condition_3][columnas_hoja3])
    df_hoja3.to_excel(writer, sheet_name='Nuevos Productos', index=False)

    writer.save()

    
#
def adjust_columns_and_center_text(path):
    try:
        # Obtenemos la lista de archivos en el directorio
        files = os.listdir(path)
        print(f"Archivos en el directorio {path}: {files}")

        for file in files:
            # Procesamos solo los archivos .xlsx
            if file.endswith('.xlsx'):
                print(f"Procesando archivo: {file}")
                try:
                    # Abrimos el archivo usando openpyxl
                    workbook = px.load_workbook(os.path.join(path, file))

                    # Iteramos a través de todas las hojas en el archivo
                    for sheet_name in workbook.sheetnames:
                        sheet = workbook[sheet_name]

                        # Iteramos a través de todas las columnas en la hoja
                        for col in sheet.columns:
                            max_length = 0
                            column = col[0].column_letter  # Obtenemos la letra de la columna

                            # Encontramos el máximo ancho de la columna
                            for cell in col:
                                try:
                                    max_length = max(max_length, len(str(cell.value)))
                                    # Centramos el texto en la celda
                                    cell.alignment = px.styles.Alignment(horizontal='center')
                                except:
                                    print("Error al ajustar celda:", cell)
                                    pass

                            # Ajustamos el ancho de la columna
                            adjusted_width = (max_length + 2)
                            sheet.column_dimensions[column].width = adjusted_width

                    # Guardamos los cambios en el archivo
                    workbook.save(os.path.join(path, file))
                    print(f"Archivo ajustado y centrado: {file}")
                except Exception as e:
                    print("Error en el archivo:", file)
                    traceback.print_exc()
    except Exception as e:
        print("Error en adjust_columns_and_center_text:")
        traceback.print_exc()

#     
def process_files(current_user_id, proveedor, path):
    # 1. Renombrar archivos
    rename_files(current_user_id, proveedor, path)
    # Inicializar los dataframes como None
    df_canal = None
    df_proveedor = None
    # Leer y procesar los archivos
    files = os.listdir(path)
    print(f"Lista de archivos después de renombrar: {files}")

    for file in files:
        if file.endswith('.xls') or file.endswith('.xlsx'):
            print(f"Procesando archivo: {file}")
            try:
                if 'stock' in file or 'precios' in file:
                    df, filename = leer_archivo(os.path.join(path, file), skiprows=0, sku_as_str='precios' in file)
                else:
                    df, filename = leer_archivo(os.path.join(path, file))
                # 3. Eliminar formas no imágenes
                eliminar_formas_no_imagenes(filename)
                # 4. Filtrar y reformatear el DataFrame
                if 'canal.xlsx' in os.path.basename(filename):
                    df_canal = filtrar_y_reformatear_canal(df)
                    df_canal.to_excel(filename, index=False)
                # Procesar archivos de proveedor
                if '-proveedor.xlsx' in os.path.basename(filename):
                    # Importar el script del proveedor
                    script_proveedor = importlib.import_module(f'proveedores.{current_user_id}.{proveedor.id}')
                    # Añade la impresión del DataFrame aquí antes de la llamada a process_proveedor_file
                    print(f"DataFrame antes de process_proveedor_file para el archivo {file}:")
                    print(df.head(10))
                    # Llamar a la función process_proveedor_file en el script del proveedor
                    df_proveedor = script_proveedor.process_proveedor_file(df)  # Cambiado a df_proveedor
                    # Añade la impresión del DataFrame aquí
                    print(f"DataFrame después de process_proveedor_file para el archivo {file}:")
                    print(df_proveedor.head(10))
                    df_proveedor.to_excel(filename, index=False)  # Cambiado a df_proveedor
            except Exception as e:
                print(f"Error al procesar el archivo {file}:")
                traceback.print_exc()

    # Combinar el dataframe combinado con el dataframe de proveedor
    if df_canal is not None and df_proveedor is not None:
        
        df_final = combine_dataframes(df_canal, df_proveedor)
        # Llamar a la función para generar el informe
        # generate_report_and_pdf(df_final, path, proveedor.nombre)
        df_final.to_excel(os.path.join(path, "final.xlsx"), index=False)
        # Crear un nuevo archivo Excel con las hojas y columnas seleccionadas
        seleccionar_columnas(df_final, proveedor, path)
        adjust_columns_and_center_text(path)
    else:
        print("No se encontraron los archivos combinados y de proveedor.")
